import csv
from openpyxl import Workbook
from openpyxl.styles import Font, DEFAULT_FONT
from openpyxl.styles.alignment import Alignment
from openpyxl.drawing.image import Image




DEFAULT_FONT.name = 'Times'
DEFAULT_FONT.size = 13

def generate_Marksheet(positive,negative):
  marksforcorrectanswer=positive
  marksforincorrectanswer=negative
  answer=[]
  j=-1
  z=0
  fstoccur=-1
  check=-1
  with open('./sample_input/responses.csv','r') as f:
     reader1=list(csv.reader(f))
     for row1 in reader1:
        fstoccur=fstoccur+1
        if fstoccur==0:
           continue
        wb=Workbook()
        sheet=wb.active
      #   sheet.column_dimensions['A'].width = 50
        img = Image('./logo.jpeg')
        img.anchor = 'A1'
        sheet.add_image(img)

        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 15

        sheet.cell(row=5,column=3).value='Mark Sheet'
        sheet.cell(row=5,column=3).font=Font(bold=True,underline='single',size=18)
        
        sheet.cell(row=6,column=1).value='Name:'
        sheet.cell(row=6,column=2).value=row1[3]
        sheet.cell(row=6,column=2).font=Font(bold=True)

        sheet.cell(row=7,column=1).value='Roll Number:'
        sheet.cell(row=7,column=2).value=row1[6]
        sheet.cell(row=7,column=2).font=Font(bold=True)

        sheet.cell(row=6,column=4).value='Exam:'

        sheet.cell(row=6,column=5).value='quiz'
        sheet.cell(row=6,column=5).font=Font(bold=True)

        sheet.cell(row=9,column=2).value='Right'  
        sheet.cell(row=9,column=2).font= Font(bold=True)
        sheet.cell(row=9,column=2).alignment=Alignment(horizontal='center')

        sheet.cell(row=9,column=3).value='Wrong'
        sheet.cell(row=9,column=3).font=Font(bold=True)
        sheet.cell(row=9,column=3).alignment=Alignment(horizontal='center')

        sheet.cell(row=9,column=4).value='Not Attempt'
        sheet.cell(row=9,column=4).font=Font(bold=True)
        sheet.cell(row=9,column=4).alignment=Alignment(horizontal='center')

        sheet.cell(row=9,column=5).value='Max'
        sheet.cell(row=9,column=5).font=Font(bold=True)
        sheet.cell(row=9,column=5).alignment=Alignment(horizontal='center')

        sheet.cell(row=10,column=1).value='No.'
        sheet.cell(row=10,column=1).font=Font(bold=True)
        sheet.cell(row=10,column=1).alignment=Alignment(horizontal='center')

        sheet.cell(row=11,column=1).value='Marking'
        sheet.cell(row=11,column=1).font=Font(bold=True)
        sheet.cell(row=11,column=1).alignment=Alignment(horizontal='center')

        sheet.cell(row=11,column=2).value=marksforcorrectanswer
        sheet.cell(row=11,column=2).font=Font(color='008000')
        sheet.cell(row=11,column=2).alignment=Alignment(horizontal='center')

        sheet.cell(row=11,column=3).value=marksforincorrectanswer
        sheet.cell(row=11,column=3).font=Font(color='FF0000')
        sheet.cell(row=11,column=3).alignment=Alignment(horizontal='center')

        sheet.cell(row=11,column=4).value=0
        sheet.cell(row=111,column=4).alignment=Alignment(horizontal='center')
        sheet.cell(row=11,column=4).alignment=Alignment(horizontal='center')

        sheet.cell(row=12,column=1).value='Total'
        sheet.cell(row=12,column=1).font=Font(bold=True)
        sheet.cell(row=12,column=1).alignment=Alignment(horizontal='center')

        sheet.cell(row=15,column=1).value='Student Ans'
        sheet.cell(row=15,column=1).font=Font(bold=True)
        sheet.cell(row=15,column=1).alignment=Alignment(horizontal='center')

        sheet.cell(row=15,column=2).value='Correct Ans'
        sheet.cell(row=15,column=2).font=Font(bold=True)
        sheet.cell(row=15,column=2).alignment=Alignment(horizontal='center')

        sheet.cell(row=15,column=4).value='Student Ans'
        sheet.cell(row=15,column=4).font=Font(bold=True)
        sheet.cell(row=15,column=4).alignment=Alignment(horizontal='center')

        sheet.cell(row=15,column=5).value='Correct Ans'
        sheet.cell(row=15,column=5).font=Font(bold=True)
        sheet.cell(row=15,column=5).alignment=Alignment(horizontal='center')

        total_count=0
        correct_count=0
        incorrect_count=0
        if row1[6]=='ANSWER':
           answer = row1
        x=16
        y=2
        for i in range(len(answer)):
           if x>40:
                 x=16
                 y=5
           if i>6:
              sheet.cell(row=x,column=y).value=answer[i]
              sheet.cell(row=x,column=y).font=Font(color='0000FF')
              sheet.cell(row=x,column=y).alignment=Alignment(horizontal='center')
              total_count=total_count+1
              x=x+1
           z=total_count  

        a=16
        b=1
        for col in range(len(row1)):
          j=j+1
          if j==0:
             continue
          if a>40:
                   a=16
                   b=4
          if col>6:
                 sheet.cell(row=a,column=b).value=row1[col]
                 if sheet.cell(row=a,column=b).value==sheet.cell(row=a,column=b+1).value:
                    sheet.cell(row=a,column=b).font=Font(color='008000')
                    sheet.cell(row=a,column=b).alignment=Alignment(horizontal='center')
                    correct_count=correct_count+1
                 elif sheet.cell(row=a,column=b).value!=sheet.cell(row=a,column=b+1).value and sheet.cell(row=a,column=b).value!='':
                    sheet.cell(row=a,column=b).font=Font(color='FF0000')
                    sheet.cell(row=a,column=b).alignment=Alignment(horizontal='center')
                    incorrect_count=incorrect_count+1
                 a=a+1
       
        sheet.cell(row=10,column=5).value=z
        sheet.cell(row=10,column=5).alignment=Alignment(horizontal='center')

        sheet.cell(row=10,column=2).value=correct_count
        sheet.cell(row=10,column=2).font=Font(color='008000')
        sheet.cell(row=10,column=2).alignment=Alignment(horizontal='center')

        sheet.cell(row=10,column=3).value=incorrect_count
        sheet.cell(row=10,column=3).font=Font(color='FF0000')
        sheet.cell(row=10,column=3).alignment=Alignment(horizontal='center')

        sheet.cell(row=10,column=4).value=z-(correct_count+incorrect_count)
        sheet.cell(row=10,column=4).alignment=Alignment(horizontal='center')

        totalmarks=z*marksforcorrectanswer

        correctmarks=marksforcorrectanswer*correct_count

        wrongmarks=marksforincorrectanswer*incorrect_count
        
        sheet.cell(row=12,column=2).value=correctmarks
        sheet.cell(row=12,column=2).font=Font(color='008000')
        sheet.cell(row=12,column=2).alignment=Alignment(horizontal='center')

        sheet.cell(row=12,column=3).value=wrongmarks
        sheet.cell(row=12,column=3).font=Font(color='FF0000')
        sheet.cell(row=12,column=3).alignment=Alignment(horizontal='center')

        got=correctmarks+wrongmarks
        sheet.cell(row=12,column=5).value=str(got)+'/'+str(totalmarks)
        sheet.cell(row=12,column=5).alignment=Alignment(horizontal='center')

        wb.save(r'my output/'+row1[6]+'.xlsx')


#generate_Marksheet(5,-1)