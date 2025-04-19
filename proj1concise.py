import csv
from openpyxl import Workbook



def concise_marksheet(positive,negative):
   marksforcorrectanswer=positive
   marksforincorrectanswer=negative
   mylist=[]
   f= open('./sample_input/responses.csv','r')
   with f:
      reader1=csv.reader(f)
      a=0
      wb=Workbook()
      sheet=wb.active
      for row1 in reader1:
         a=a+1
         b=1
         for i in range(len(row1)):
            if row1[6]=='ANSWER':
               if i>6:
                  mylist.append(row1[i])
            if b==7:
             b=8
            sheet.cell(row=a,column=b).value=row1[i]
            b=b+1
      k=-1
      d=1
      f1= open('./sample_input/responses.csv','r')
      with f1:
       reader2=csv.reader(f1)
       for row2 in reader2:
          index=0
          correctval=0
          incorrectval=0
          k=k+1
          if k==0:
             continue
            
          for j in range(len(row2)):
             if j>6:
                if row2[j]==mylist[index]:
                   correctval=correctval+1
                elif row2[j]!=mylist[index] and row2[j]!='':
                   incorrectval=incorrectval+1
                index=index+1
          d=d+1
          sheet.cell(row=d,column=7).value=str((correctval*marksforcorrectanswer)+(incorrectval*marksforincorrectanswer))+'/'+str(len(mylist)*marksforcorrectanswer)
          sheet.cell(row=d,column=37).value=str([correctval,incorrectval,correctval-incorrectval])
      sheet.cell(row=1,column=3).value='Google_score'
      sheet.cell(row=1,column=7).value='score_after'
   
      

      wb.save(r'my output/'+'concise_marksheet'+'.xlsx')   